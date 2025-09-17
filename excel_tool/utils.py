import io
import re
import pandas as pd
import openpyxl
from datetime import datetime
from copy import copy
from typing import Dict, Any, Optional
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

NUMERIC_TOLERANCE = 1

def coerce_numeric(value):
    if pd.isna(value) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s == "":
        return None
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    s = s.replace(",", "")
    s = re.sub(r"[^\d\.\-]+", "", s)
    try:
        return float(s)
    except Exception:
        return None

def to_int_or_none(value):
    num = coerce_numeric(value)
    if num is None:
        return None
    try:
        return int(num)
    except Exception:
        return None

def norm_text(value):
    if pd.isna(value) or value is None:
        return ""
    return str(value).strip()

def values_equal(orig_val, web_val, tol=NUMERIC_TOLERANCE):
    oi = to_int_or_none(orig_val)
    wi = to_int_or_none(web_val)
    if oi is not None and wi is not None:
        return abs(wi - oi) <= tol
    return norm_text(orig_val) == norm_text(web_val)

def values_different(orig_val, web_val, tol=NUMERIC_TOLERANCE):
    return not values_equal(orig_val, web_val, tol=tol)

def row_key(row, max_parts=2):
    key_parts = []
    for col in row.index:
        iv = to_int_or_none(row[col])
        if iv is not None:
            key_parts.append(("num", iv))
        else:
            sv = norm_text(row[col])
            if sv != "":
                key_parts.append(("str", sv.lower()))
        if len(key_parts) >= max_parts:
            break
    return tuple(key_parts)

def build_key_index(df, max_parts=2):
    idx = {}
    for i, row in df.iterrows():
        k = row_key(row, max_parts=max_parts)
        idx.setdefault(k, []).append(i)
    return idx

def safe_copy_cell_style(source_cell, target_cell):
    try:
        if source_cell is None or target_cell is None:
            return
        if hasattr(source_cell, "has_style") and source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
    except Exception:
        pass

def safe_get_column_width(worksheet, col_letter):
    try:
        if col_letter in worksheet.column_dimensions:
            return worksheet.column_dimensions[col_letter].width
        return None
    except Exception:
        return None

def norm_for_json(v):
    if pd.isna(v) or v is None:
        return ""
    if isinstance(v, (int, float, bool)):
        try:
            return int(v) if isinstance(v, int) and not isinstance(v, bool) else float(v) if isinstance(v, float) else v
        except Exception:
            return v
    return str(v)

# Main compare function
def compare_excel_with_gain_summary_inline(
    original_bytes: bytes,
    website_bytes: bytes,
    sheets_config: Optional[Dict[str, Dict[str, int]]] = None,
) -> Dict[str, Any]:

    if sheets_config is None:
        sheets_config = {
            'Gain Summary': {'header_row': 2, 'data_start_row': 3},
            '8938':         {'header_row': 6, 'data_start_row': 7},
            'FBAR':         {'header_row': 2, 'data_start_row': 3},
        }

    red_fill   = PatternFill(start_color="fbd9d3", end_color="fbd9d3", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    original_wb = openpyxl.load_workbook(io.BytesIO(original_bytes))
    website_wb  = openpyxl.load_workbook(io.BytesIO(website_bytes))
    output_wb   = openpyxl.Workbook()
    if 'Sheet' in output_wb.sheetnames:
        output_wb.remove(output_wb['Sheet'])

    summary_rows = [["Sheet", "Rows Compared", "Cells Different", "Common Rows", "Only in Original", "Only in Website"]]
    sheets_structured: Dict[str, Dict[str, Any]] = {}
    out_bytes = io.BytesIO()
    output_wb.save(out_bytes)
    out_bytes.seek(0)

    return {
        "excel_bytes": out_bytes.getvalue(),
        "summary_rows": summary_rows,
        "sheets": sheets_structured
    }

def write_html_report(summary_rows):
    html = """
    <html>
    <head>
        <title>Excel Comparison Report</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; }
            table { border-collapse: collapse; width: 100%; }
            th, td { border: 1px solid #ccc; padding: 6px; text-align: left; }
            th { background: #f2f2f2; }
        </style>
    </head>
    <body>
        <h1>Excel Comparison Report</h1>
        <table>
            <tr>
                <th>Sheet</th>
                <th>Rows Compared</th>
                <th>Cells Different</th>
                <th>Common Rows</th>
                <th>Only in Original</th>
                <th>Only in Website</th>
            </tr>
    """
    for row in summary_rows[1:]:
        html += "<tr>" + "".join(f"<td>{cell}</td>" for cell in row) + "</tr>"
    html += """
        </table>
    </body>
    </html>
    """
    return html