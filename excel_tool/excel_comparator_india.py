import io
import re
from copy import copy
from datetime import datetime
from typing import Dict, Any, Optional
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import traceback  # added import for detailed exception traces


NUMERIC_TOLERANCE = 1


def coerce_numeric(value):
    if pd.isna(value) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s == "":
        return None
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    s = s.replace(",", "")
    s = re.sub(r"[^\d\.\-]+", "", s)
    try:
        return float(s)
    except Exception:
        return None


def to_int_or_none(value):
    num = coerce_numeric(value)
    if num is None:
        return None
    try:
        return int(num)
    except Exception:
        return None


def norm_text(value):
    if pd.isna(value) or value is None:
        return ""
    return str(value).strip()


def values_equal(orig_val, web_val, tol=NUMERIC_TOLERANCE):
    oi = to_int_or_none(orig_val)
    wi = to_int_or_none(web_val)
    if oi is not None and wi is not None:
        return abs(wi - oi) <= tol
    return norm_text(orig_val) == norm_text(web_val)


def values_different(orig_val, web_val, tol=NUMERIC_TOLERANCE):
    return not values_equal(orig_val, web_val, tol=tol)


def row_key(row, max_parts=2):
    key_parts = []
    for col in row.index:
        iv = to_int_or_none(row[col])
        if iv is not None:
            key_parts.append(("num", iv))
        else:
            sv = norm_text(row[col])
            if sv != "":
                key_parts.append(("str", sv.lower()))
        if len(key_parts) >= max_parts:
            break
    return tuple(key_parts)


def build_key_index(df, max_parts=2):
    idx = {}
    for i, row in df.iterrows():
        k = row_key(row, max_parts=max_parts)
        idx.setdefault(k, []).append(i)
    return idx


def safe_copy_cell_style(source_cell, target_cell):
    try:
        if source_cell is None or target_cell is None:
            return
        if hasattr(source_cell, "has_style") and source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
    except Exception:
        pass


def safe_get_column_width(worksheet, col_letter):
    try:
        if col_letter in worksheet.column_dimensions:
            return worksheet.column_dimensions[col_letter].width
        return None
    except Exception:
        return None


def norm_for_json(v):
    if pd.isna(v) or v is None:
        return ""
    if isinstance(v, (int, float, bool)):
        try:
            return int(v) if isinstance(v, int) and not isinstance(v, bool) else float(v) if isinstance(v, float) else v
        except Exception:
            return v
    return str(v)


# New helper to ensure sheet titles are <= 31 chars and unique within workbook
def make_safe_title(title: str, wb: openpyxl.workbook.workbook.Workbook) -> str:
    if title is None:
        title = ""
    max_len = 31
    base = title[:max_len]
    safe = base
    i = 1
    # if exact title exists, attempt to append suffix like " (1)" ensuring total length <= max_len
    while safe in wb.sheetnames:
        suffix = f" ({i})"
        keep_len = max_len - len(suffix)
        safe = (base[:keep_len] + suffix) if keep_len > 0 else suffix[:max_len]
        i += 1
        if i > 999:
            break
    return safe


# Main compare function
def compare_excel_with_gain_summary_inline_India(
    original_bytes: bytes,
    website_bytes: bytes,
    sheets_config: Optional[Dict[str, Dict[str, int]]] = None,
) -> Dict[str, Any]:

    if sheets_config is None:
        sheets_config = {
            'Gain Summary': {'header_row': 2, 'data_start_row': 3},
            'ScheduleFA':         {'header_row': 8, 'data_start_row': 8},
            'transaction_details':         {'header_row': 2, 'data_start_row': 3},
            'transaction_details_by_gain':         {'header_row': 2, 'data_start_row': 3},
        }

    red_fill   = PatternFill(start_color="fbd9d3", end_color="fbd9d3", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    original_wb = openpyxl.load_workbook(io.BytesIO(original_bytes))
    website_wb  = openpyxl.load_workbook(io.BytesIO(website_bytes))
    output_wb   = openpyxl.Workbook()
    if 'Sheet' in output_wb.sheetnames:
        output_wb.remove(output_wb['Sheet'])

    summary_rows = [["Sheet", "Rows Compared", "Cells Different", "Common Rows", "Only in Original", "Only in Website"]]

    sheets_structured: Dict[str, Dict[str, Any]] = {}

    def acct_key(v):
        if pd.isna(v) or v is None:
            return None
        s = str(v).strip()
        if s == "":
            return None
        num = to_int_or_none(s)
        if num is not None:
            return str(num)
        return s.lower()

    # New helper: build composite key from Account Number, Investment Name and Purchase Date
    def composite_acct_key(row, key_cols):
        """
        Normalize and combine values from key_cols into a single string key.
        Returns None if any required part is missing/blank.
        """
        parts = []
        for col in key_cols:
            if col not in row.index:
                return None
            v = row[col]
            if pd.isna(v) or v is None:
                return None
            s = str(v).strip()
            if s == "":
                return None

            if col.lower() == "account number":
                # prefer normalized numeric account if possible
                num = to_int_or_none(v)
                parts.append(str(num) if num is not None else s.lower())
            elif "date" in col.lower():
                # normalize dates to ISO date if possible
                try:
                    ts = pd.to_datetime(v, errors="coerce", dayfirst=True)
                    if pd.isna(ts):
                        parts.append(s.lower())
                    else:
                        parts.append(ts.date().isoformat())
                except Exception:
                    parts.append(s.lower())
            else:
                parts.append(s.lower())
        return "|".join(parts)

    for sheet_name, cfg in sheets_config.items():
        try:
            df_orig = pd.read_excel(io.BytesIO(original_bytes), sheet_name=sheet_name, header=cfg['header_row'] - 1)
            df_web  = pd.read_excel(io.BytesIO(website_bytes),  sheet_name=sheet_name, header=cfg['header_row'] - 1)

            common_cols = [c for c in df_orig.columns if c in df_web.columns]

            if sheet_name in ("ScheduleFA", "transaction_details", "transaction_details_by_gain"):
                common_cols = [
                    c for c in common_cols 
                    if c != "Sl. No."
                ]

            structured_common = []
            structured_diff   = []
            structured_only_orig = []
            structured_only_web  = []

            if not common_cols:
                output_wb.create_sheet(title=make_safe_title(sheet_name, output_wb))
                output_wb.create_sheet(title=make_safe_title(f"{sheet_name} Common Rows", output_wb))
                summary_rows.append([sheet_name, 0, 0, 0, 0, 0])
                sheets_structured[sheet_name] = {
                    "common_rows": structured_common,
                    "different_rows": structured_diff,
                    "only_in_original_rows": structured_only_orig,
                    "only_in_website_rows": structured_only_web,
                    "rows_compared": 0,
                    "cells_different": 0
                }
                continue

            df_orig = df_orig[common_cols].copy()
            df_web  = df_web[common_cols].copy()

            original_ws = original_wb[sheet_name] if sheet_name in original_wb.sheetnames else None
            website_ws  = website_wb[sheet_name]  if sheet_name in website_wb.sheetnames  else None
            main_title = make_safe_title(sheet_name, output_wb)
            main_ws     = output_wb.create_sheet(title=main_title)

            diff_count       = 0
            rows_compared    = 0
            common_rows_list = []

            def make_tripled_headers():
                headers = []
                for col in common_cols:
                    headers.extend([f"{col} (Original)", f"{col} (Website)", f"{col} (Diff)"])
                structured_common.append([h for h in headers])
                structured_diff.append([h for h in headers])
                structured_only_orig.append([h for h in headers])
                structured_only_web.append([h for h in headers])
                for ci, h in enumerate(headers, start=1):
                    main_ws.cell(row=1, column=ci, value=h)
                    if original_ws is not None:
                        try:
                            src = original_ws.cell(row=cfg['header_row'], column=((ci - 1) // 3) + 1)
                            safe_copy_cell_style(src, main_ws.cell(row=1, column=ci))
                        except Exception:
                            traceback.print_exc()

                for idx_col, _ in enumerate(common_cols, start=1):
                    width = None
                    if original_ws is not None:
                        width = safe_get_column_width(original_ws, get_column_letter(idx_col))
                    for offset in range(3):
                        out_col_letter = get_column_letter((idx_col - 1) * 3 + offset + 1)
                        if width:
                            main_ws.column_dimensions[out_col_letter].width = width

            def make_simple_headers():
                headers = [c for c in common_cols] + ["Row Status"]
                structured_common.append([h for h in headers])
                structured_diff.append([h for h in headers])
                structured_only_orig.append([h for h in headers])
                structured_only_web.append([h for h in headers])

                for ci, h in enumerate(headers, start=1):
                    main_ws.cell(row=1, column=ci, value=h)
                    if original_ws is not None:
                        try:
                            src = original_ws.cell(row=cfg['header_row'], column=ci)
                            safe_copy_cell_style(src, main_ws.cell(row=1, column=ci))
                        except Exception:
                            traceback.print_exc()

                main_ws.column_dimensions[get_column_letter(len(headers))].width = 18

            if sheet_name in ("Gain Summary", "ScheduleFA", "transaction_details","transaction_details_by_gain") and ("Account Number" in df_orig.columns) and ("Account Number" in df_web.columns):
                make_tripled_headers()

                # For transaction detail sheets prefer using Account Number + Investment Name + Purchase Date
                preferred_keys = ["Account Number", "Investment Name", "Purchase Date"]
                use_composite = all(k in df_orig.columns and k in df_web.columns for k in preferred_keys)
                if use_composite:
                    key_cols = preferred_keys
                else:
                    key_cols = ["Account Number"]

                # Deduplicate using chosen key columns where possible
                try:
                    df_orig = df_orig.drop_duplicates(subset=key_cols, keep="first").reset_index(drop=True)
                    df_web  = df_web.drop_duplicates(subset=key_cols, keep="first").reset_index(drop=True)
                except Exception:
                    # fallback if drop_duplicates fails for any reason
                    df_orig = df_orig.reset_index(drop=True)
                    df_web = df_web.reset_index(drop=True)

                # Gain Summary
                if sheet_name == "Gain Summary":
                    web_key_index = build_key_index(df_web, max_parts=2)
                    orig_keys_set = {row_key(r, max_parts=2) for _, r in df_orig.iterrows()}

                    out_row = 2
                    for orig_idx, orig_row in df_orig.iterrows():
                        k = row_key(orig_row, max_parts=2)
                        match_indices = web_key_index.get(k, [])
                        match_idx = match_indices[0] if match_indices else None

                        if match_idx is None:
                            triple_row = []
                            for col in common_cols:
                                o_val = norm_for_json(orig_row[col])
                                triple_row.extend([o_val, "", "Only in Original"])
                                main_ws.cell(row=out_row, column=((common_cols.index(col)) * 3) + 1, value=o_val)
                                main_ws.cell(row=out_row, column=((common_cols.index(col)) * 3) + 3, value="Only in Original").fill = red_fill
                            structured_only_orig.append(triple_row)
                            out_row += 1
                            rows_compared += 1
                            diff_count += 1
                            continue

                        web_row = df_web.loc[match_idx]
                        all_same = all(values_equal(orig_row[col], web_row[col]) for col in common_cols)
                        if all_same:
                            triple_row = []
                            for col in common_cols:
                                val = norm_for_json(orig_row[col])
                                triple_row.extend([val, val, ""])
                                ci = common_cols.index(col)
                                main_ws.cell(row=out_row, column=ci * 3 + 1, value=orig_row[col])
                                main_ws.cell(row=out_row, column=ci * 3 + 2, value=web_row[col])
                            structured_common.append(triple_row)
                            out_row += 1
                            continue

                        row_diffs_here = 0
                        triple_row = []
                        for col in common_cols:
                            o_val = orig_row[col]
                            w_val = web_row[col]
                            is_diff = values_different(o_val, w_val)
                            diff_val = ""
                            if is_diff:
                                o_num = coerce_numeric(o_val)
                                w_num = coerce_numeric(w_val)
                                if o_num is not None and w_num is not None:
                                    diff_val = w_num - o_num
                                else:
                                    diff_val = "DIFF"
                                row_diffs_here += 1

                            triple_row.extend([norm_for_json(o_val), norm_for_json(w_val), norm_for_json(diff_val)])
                            ci = common_cols.index(col)
                            c_orig = main_ws.cell(row=out_row, column=ci * 3 + 1, value=o_val)
                            c_web  = main_ws.cell(row=out_row, column=ci * 3 + 2, value=w_val)
                            c_diff = main_ws.cell(row=out_row, column=ci * 3 + 3, value=diff_val)
                            if is_diff:
                                c_orig.fill = red_fill
                                c_web.fill  = green_fill
                                c_diff.fill = red_fill

                        structured_diff.append(triple_row)
                        out_row += 1
                        rows_compared += 1
                        diff_count += row_diffs_here

                    for web_idx, web_row in df_web.iterrows():
                        k = row_key(web_row, max_parts=2)
                        if k in orig_keys_set:
                            continue
                        triple_row = []
                        for col in common_cols:
                            w_val = norm_for_json(web_row[col])
                            triple_row.extend(["", w_val, "Only in Website"])
                        structured_only_web.append(triple_row)
                        for col in common_cols:
                            ci = common_cols.index(col)
                            main_ws.cell(row=out_row, column=ci * 3 + 2, value=web_row[col])
                            main_ws.cell(row=out_row, column=ci * 3 + 3, value="Only in Website").fill = green_fill
                        out_row += 1
                        rows_compared += 1
                        diff_count += 1

                else:
                    # build maps using composite or single key as selected above
                    orig_map = {}
                    web_map = {}
                    if len(key_cols) == 1 and key_cols[0] == "Account Number":
                        for i, r in df_orig.iterrows():
                            k = acct_key(r["Account Number"])
                            if k is not None and k not in orig_map:
                                orig_map[k] = i
                        for i, r in df_web.iterrows():
                            k = acct_key(r["Account Number"])
                            if k is not None and k not in web_map:
                                web_map[k] = i
                    else:
                        for i, r in df_orig.iterrows():
                            k = composite_acct_key(r, key_cols)
                            if k is not None and k not in orig_map:
                                orig_map[k] = i
                        for i, r in df_web.iterrows():
                            k = composite_acct_key(r, key_cols)
                            if k is not None and k not in web_map:
                                web_map[k] = i

                    out_row = 2
                    for acc in orig_map.keys():
                        orig_idx = orig_map[acc]
                        orig_row = df_orig.loc[orig_idx]
                        if acc not in web_map:
                            triple_row = []
                            for col in common_cols:
                                o_val = norm_for_json(orig_row[col])
                                triple_row.extend([o_val, "", "Only in Original"])
                                ci = common_cols.index(col)
                                main_ws.cell(row=out_row, column=ci * 3 + 1, value=orig_row[col])
                                main_ws.cell(row=out_row, column=ci * 3 + 3, value="Only in Original").fill = red_fill
                            structured_only_orig.append(triple_row)
                            out_row += 1
                            rows_compared += 1
                            diff_count += 1
                            continue

                        web_idx = web_map[acc]
                        web_row = df_web.loc[web_idx]
                        all_same = all(values_equal(orig_row[col], web_row[col]) for col in common_cols)
                        if all_same:
                            triple_row = []
                            for col in common_cols:
                                val = norm_for_json(orig_row[col])
                                triple_row.extend([val, val, ""])
                                ci = common_cols.index(col)
                                main_ws.cell(row=out_row, column=ci * 3 + 1, value=orig_row[col])
                                main_ws.cell(row=out_row, column=ci * 3 + 2, value=web_row[col])
                            structured_common.append(triple_row)
                            out_row += 1
                            continue

                        row_diffs_here = 0
                        triple_row = []
                        for col in common_cols:
                            o_val = orig_row[col]
                            w_val = web_row[col]
                            is_diff = values_different(o_val, w_val)
                            diff_val = ""
                            if is_diff:
                                o_num = coerce_numeric(o_val)
                                w_num = coerce_numeric(w_val)
                                if o_num is not None and w_num is not None:
                                    diff_val = w_num - o_num
                                else:
                                    diff_val = "DIFF"
                                row_diffs_here += 1

                            triple_row.extend([norm_for_json(o_val), norm_for_json(w_val), norm_for_json(diff_val)])
                            ci = common_cols.index(col)
                            c_orig = main_ws.cell(row=out_row, column=ci * 3 + 1, value=o_val)
                            c_web  = main_ws.cell(row=out_row, column=ci * 3 + 2, value=w_val)
                            c_diff = main_ws.cell(row=out_row, column=ci * 3 + 3, value=diff_val)
                            if is_diff:
                                c_orig.fill = red_fill
                                c_web.fill  = green_fill
                                c_diff.fill = red_fill

                        structured_diff.append(triple_row)
                        out_row += 1
                        rows_compared += 1
                        diff_count += row_diffs_here

                    for acc in web_map.keys():
                        if acc in orig_map:
                            continue
                        web_idx = web_map[acc]
                        web_row = df_web.loc[web_idx]
                        triple_row = []
                        for col in common_cols:
                            w_val = norm_for_json(web_row[col])
                            triple_row.extend(["", w_val, "Only in Website"])
                            ci = common_cols.index(col)
                            main_ws.cell(row=out_row, column=ci * 3 + 2, value=web_row[col])
                            main_ws.cell(row=out_row, column=ci * 3 + 3, value="Only in Website").fill = green_fill
                        structured_only_web.append(triple_row)
                        out_row += 1
                        rows_compared += 1
                        diff_count += 1

            # General compare logic
            else:
                make_simple_headers()
                web_key_index = build_key_index(df_web, max_parts=2)
                out_row = 2

                orig_key_set = {row_key(r, max_parts=2) for _, r in df_orig.iterrows()}

                for orig_idx, orig_row in df_orig.iterrows():
                    k = row_key(orig_row, max_parts=2)
                    match_indices = web_key_index.get(k, [])
                    match_idx = match_indices[0] if match_indices else None

                    if match_idx is None:
                        rowvals = [norm_for_json(orig_row[col]) for col in common_cols] + ["Only in Original"]
                        structured_only_orig.append(rowvals)
                        for i, col in enumerate(common_cols, start=1):
                            main_ws.cell(row=out_row, column=i, value=orig_row[col])
                        main_ws.cell(row=out_row, column=len(common_cols) + 1, value="Only in Original").fill = red_fill
                        out_row += 1
                        rows_compared += 1
                        diff_count += 1
                        continue

                    web_row = df_web.loc[match_idx]
                    if all(values_equal(orig_row[col], web_row[col]) for col in common_cols):
                        rowvals = [norm_for_json(orig_row[col]) for col in common_cols] + ["Common"]
                        structured_common.append(rowvals)
                        for i, col in enumerate(common_cols, start=1):
                            main_ws.cell(row=out_row, column=i, value=orig_row[col])
                        out_row += 1
                        continue

                    row_diffs_here = 0
                    rowvals = []
                    for col in common_cols:
                        o_val = orig_row[col]
                        w_val = web_row[col]
                        if values_different(o_val, w_val):
                            row_diffs_here += 1
                        rowvals.append(norm_for_json(o_val))
                    rowvals.append("Different")
                    structured_diff.append(rowvals)
                    for i, col in enumerate(common_cols, start=1):
                        c = main_ws.cell(row=out_row, column=i, value=orig_row[col])
                        if values_different(orig_row[col], web_row[col]):
                            c.fill = red_fill
                    main_ws.cell(row=out_row, column=len(common_cols) + 1, value="Different").fill = red_fill
                    out_row += 1
                    rows_compared += 1
                    diff_count += row_diffs_here

                for web_idx, web_row in df_web.iterrows():
                    k = row_key(web_row, max_parts=2)
                    if k in orig_key_set:
                        continue
                    rowvals = [norm_for_json(web_row[col]) for col in common_cols] + ["Only in Website"]
                    structured_only_web.append(rowvals)
                    for i, col in enumerate(common_cols, start=1):
                        main_ws.cell(row=out_row, column=i, value=web_row[col])
                    main_ws.cell(row=out_row, column=len(common_cols) + 1, value="Only in Website").fill = green_fill
                    out_row += 1
                    rows_compared += 1
                    diff_count += 1

            for ws_auto in (main_ws,):
                for col in ws_auto.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value is not None:
                            try:
                                max_len = max(max_len, len(str(cell.value)))
                            except Exception:
                                max_len = max(max_len, 10)
                    ws_auto.column_dimensions[col_letter].width = min(max_len + 2, 50)

            sheets_structured[sheet_name] = {
                "common_rows": structured_common,
                "different_rows": structured_diff,
                "only_in_original_rows": structured_only_orig,
                "only_in_website_rows": structured_only_web,
                "rows_compared": rows_compared,
                "cells_different": diff_count
                , "_main_ws_name": main_title   # preserve the actual styled sheet name for later reuse
            }

            summary_rows.append([sheet_name, rows_compared, diff_count, len(structured_common), len(structured_only_orig), len(structured_only_web)])

        except Exception as e:
            traceback.print_exc()
            summary_rows.append([sheet_name, 0, f"ERROR: {e}", 0, 0, 0])
            sheets_structured[sheet_name] = {
                "common_rows": [],
                "different_rows": [],
                "only_in_original_rows": [],
                "only_in_website_rows": [],
                "rows_compared": 0,
                "cells_different": 0,
                "error": str(e)
            }

    for sheet_name, sdata in list(sheets_structured.items()):
        if not isinstance(sdata, dict):
            continue

        common_rows = sdata.get("common_rows") or []
        if len(common_rows) > 1:
            safe_common_title = make_safe_title(f"{sheet_name} Common Rows", output_wb)
            ws_common = output_wb.create_sheet(title=safe_common_title)
            for r in common_rows:
                ws_common.append(r)

        header = None
        for key in ("different_rows", "common_rows", "only_in_original_rows", "only_in_website_rows"):
            lst = sdata.get(key) or []
            if len(lst) > 0:
                header = lst[0]
                break

        remaining_rows = []
        for key in ("different_rows", "only_in_original_rows", "only_in_website_rows"):
            lst = sdata.get(key) or []
            if len(lst) > 1:
                remaining_rows.extend(lst[1:])

        # For transaction detail sheets, create dedicated sheets for each category (include header)
        if sheet_name in ("transaction_details", "transaction_details_by_gain"):
            for key, label in (("different_rows", "Different Rows"), ("only_in_original_rows", "Only in Original"), ("only_in_website_rows", "Only in Website")):
                lst = sdata.get(key) or []
                if len(lst) > 0:
                    cat_title = make_safe_title(f"{sheet_name} {label}", output_wb)
                    if cat_title in output_wb.sheetnames:
                        try:
                            del output_wb[cat_title]
                        except Exception:
                            pass
                    ws_cat = output_wb.create_sheet(title=cat_title)
                    for r in lst:
                        ws_cat.append(r)
                    # auto-size columns
                    for col_idx in range(1, ws_cat.max_column + 1):
                        col_letter = get_column_letter(col_idx)
                        max_len = 0
                        for cell in ws_cat[col_letter]:
                            if cell.value is not None:
                                try:
                                    max_len = max(max_len, len(str(cell.value)))
                                except Exception:
                                    max_len = max(max_len, 10)
                        ws_cat.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # If we previously created a styled "main" worksheet for this sheet, reuse it (preserve fills/styles).
        main_ws_name = sdata.get("_main_ws_name")
        if main_ws_name and main_ws_name in output_wb.sheetnames:
            new_ws = output_wb[main_ws_name]
            # No need to re-append rows — the styled sheet already contains the detailed output.
        else:
            # create new sheet with safe name (no styled sheet existed)
            safe_name = make_safe_title(sheet_name, output_wb)
            if safe_name in output_wb.sheetnames:
                try:
                    del output_wb[safe_name]
                except Exception:
                    pass
            new_ws = output_wb.create_sheet(title=safe_name)
            if header:
                new_ws.append(header)
            for r in remaining_rows:
                new_ws.append(r)
            for col_idx in range(1, new_ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for cell in new_ws[col_letter]:
                    if cell.value is not None:
                        try:
                            max_len = max(max_len, len(str(cell.value)))
                        except Exception:
                            max_len = max(max_len, 10)
                new_ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    ws_summary = output_wb.create_sheet(make_safe_title("Summary", output_wb), index=0)
    ws_summary.append(["Excel Comparison Report"])
    ws_summary.append([f"Generated On:  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws_summary.append([])
    for row in summary_rows:
        ws_summary.append(row)

    out_bytes = io.BytesIO()
    output_wb.save(out_bytes)
    out_bytes.seek(0)

    return {
        "excel_bytes": out_bytes.getvalue(),
        "summary_rows": summary_rows,
        "sheets": sheets_structured
    }


# HTML report
def write_html_report_India(summary_rows):
    html = """
    <html>
    <head>
        <title>Excel Comparison Report</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; }
            table { border-collapse: collapse; width: 100%; }
            th, td { border: 1px solid #ccc; padding: 6px; text-align: left; }
            th { background: #f2f2f2; }
        </style>
    </head>
    <body>
        <h1>Excel Comparison Report</h1>
        <table>
            <tr>
                <th>Sheet</th>
                <th>Rows Compared</th>
                <th>Cells Different</th>
                <th>Common Rows</th>
                <th>Only in Original</th>
                <th>Only in Website</th>
            </tr>
    """
    for row in summary_rows[1:]: 
        html += "<tr>" + "".join(f"<td>{cell}</td>" for cell in row) + "</tr>"

    html += """
        </table>
    </body>
    </html>
    """
    return html


def compare_excel_sheets_inline_India(original_bytes, website_bytes, sheets_config=None):
    result = compare_excel_with_gain_summary_inline_India(original_bytes, website_bytes, sheets_config)
    sheets_with_diff = []
    for sname, sdata in result["sheets"].items():
        if (len(sdata.get("different_rows", [])) > 1 or
            len(sdata.get("only_in_original_rows", [])) > 1 or
            len(sdata.get("only_in_website_rows", [])) > 1):
            sheets_with_diff.append(sname)
    return {"sheets": sheets_with_diff}


def compare_single_sheet_diff_inline_India(original_bytes, website_bytes, sheet_name, sheets_config=None):
    result = compare_excel_with_gain_summary_inline_India(original_bytes, website_bytes, sheets_config)
    sheets = result.get("sheets", {})
    target = sheets.get(sheet_name)
    if not target:
        return {
            "sheet_name": sheet_name,
            "common_rows": [],
            "different_rows": [],
            "only_in_original_rows": [],
            "only_in_website_rows": []
        }

    return {
        "sheet_name": sheet_name,
        "common_rows": target.get("common_rows", []),
        "different_rows": target.get("different_rows", []),
        "only_in_original_rows": target.get("only_in_original_rows", []),
        "only_in_website_rows": target.get("only_in_website_rows", [])
    }


def generate_comparison_excel_India(original_bytes, website_bytes, output_stream, sheets_config=None):
    result = compare_excel_with_gain_summary_inline_India(original_bytes, website_bytes, sheets_config)
    output_stream.write(result["excel_bytes"])